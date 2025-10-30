"""
Excel Export Handler

This module handles Excel file creation with tracking numbers and formatting.
Creates professional output files with auto-adjusted columns and proper styling.
"""

from datetime import datetime
from typing import List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.utils.constants import (
    COLUMN_TRACKING_NUMBER,
    COLUMN_DELIVERY_COMPANY,
    DELIVERY_COMPANY,
    ERR_EXPORT_FAILED,
)
from src.utils.validators import validate_output_path, sanitize_filename
from src.utils.logger import get_logger

# Import FileFormat from excel_uploader
from src.handlers.excel_uploader import FileFormat

logger = get_logger(__name__)


class ExcelExportError(Exception):
    """Custom exception for Excel export errors"""
    pass


class ExcelExportHandler:
    """
    Handles Excel file export with tracking numbers
    """

    @staticmethod
    def create_output(
        special_codes: List[str],
        tracking_numbers: List[str],
        output_path: str,
        apply_formatting: bool = True
    ) -> bool:
        """
        Create output Excel file with exactly 3 columns:
        1. 주문고유코드 (special code from input)
        2. 택배사 (fixed: 경동택배)
        3. 가송장 번호 (randomized tracking number)

        Args:
            special_codes: List of special codes from input file
            tracking_numbers: List of generated tracking numbers
            output_path: Path to save output file
            apply_formatting: Apply Excel formatting (default: True)

        Returns:
            bool: True if successful

        Raises:
            ExcelExportError: If export fails

        Example:
            >>> handler = ExcelExportHandler()
            >>> codes = ["DA616E9F6", "D74B2E218"]
            >>> numbers = ["20251111111111", "20252222222222"]
            >>> handler.create_output(codes, numbers, "output.xlsx")
            True
        """
        try:
            # Validate output path
            is_valid, error_message = validate_output_path(output_path)
            if not is_valid:
                raise ExcelExportError(error_message)

            # Validate input
            if len(special_codes) != len(tracking_numbers):
                raise ExcelExportError(
                    f"Mismatch: {len(special_codes)} special codes but {len(tracking_numbers)} tracking numbers provided"
                )

            # Create output DataFrame with exactly 3 columns
            output_df = pd.DataFrame({
                '주문고유코드': special_codes,
                '택배사': [DELIVERY_COMPANY] * len(special_codes),
                '가송장 번호': tracking_numbers
            })

            logger.info(f"Exporting 3-column format: {len(output_df)} rows")
            logger.info(f"Columns: {list(output_df.columns)}")

            # Write to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                output_df.to_excel(writer, index=False, sheet_name='Sheet1')

                # Apply formatting if requested
                if apply_formatting:
                    ExcelExportHandler._apply_formatting(writer, output_df)

            logger.info(f"Successfully exported to: {output_path}")
            return True

        except ExcelExportError:
            raise

        except Exception as e:
            logger.error(f"Export failed: {e}", exc_info=True)
            # Generic message to user (no internal details for security)
            raise ExcelExportError("파일을 저장할 수 없습니다. 경로를 확인하거나 다른 위치에 저장해보세요.")

    @staticmethod
    def _apply_formatting(writer, df: pd.DataFrame) -> None:
        """
        Apply professional formatting to Excel file

        Args:
            writer: pandas ExcelWriter object
            df: DataFrame being written
        """
        try:
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Header formatting
            header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # gray-100
            header_font = Font(bold=True, size=12, color="1F2937")  # gray-800
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Apply header formatting
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Format tracking number column (monospace font, centered)
            tracking_col_idx = df.columns.get_loc('가송장 번호') + 1  # +1 for Excel (1-indexed)
            tracking_col_letter = chr(64 + tracking_col_idx)  # Convert to letter

            tracking_font = Font(name="Courier New", size=11)
            tracking_alignment = Alignment(horizontal="center", vertical="center")

            for row in range(2, len(df) + 2):  # Skip header
                cell = worksheet[f"{tracking_col_letter}{row}"]
                cell.font = tracking_font
                cell.alignment = tracking_alignment

            # Format all data columns (center alignment)
            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = chr(64 + col_idx)
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            logger.debug("Applied Excel formatting")

        except Exception as e:
            logger.warning(f"Failed to apply formatting (non-critical): {e}")

    @staticmethod
    def generate_filename(prefix: str = "가송장_생성기") -> str:
        """
        Generate output filename with timestamp

        Args:
            prefix: Filename prefix (default: "가송장_생성기")

        Returns:
            str: Filename with format: prefix_YYYYMMDD_HHMMSS.xlsx

        Example:
            >>> filename = ExcelExportHandler.generate_filename()
            >>> filename.endswith('.xlsx')
            True
            >>> '가송장_생성기' in filename
            True
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{sanitize_filename(prefix)}_{timestamp}.xlsx"
        logger.debug(f"Generated filename: {filename}")
        return filename

    @staticmethod
    def append_to_existing(
        existing_path: str,
        new_data: pd.DataFrame,
        sheet_name: str = "Sheet1"
    ) -> bool:
        """
        Append data to existing Excel file (optional feature)

        Args:
            existing_path: Path to existing Excel file
            new_data: New data to append
            sheet_name: Sheet to append to (default: "Sheet1")

        Returns:
            bool: True if successful
        """
        try:
            # Read existing
            existing_df = pd.read_excel(existing_path, sheet_name=sheet_name)

            # Append new data
            combined_df = pd.concat([existing_df, new_data], ignore_index=True)

            # Write back
            with pd.ExcelWriter(existing_path, engine='openpyxl', mode='w') as writer:
                combined_df.to_excel(writer, index=False, sheet_name=sheet_name)

            logger.info(f"Appended {len(new_data)} rows to {existing_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to append data: {e}")
            return False
